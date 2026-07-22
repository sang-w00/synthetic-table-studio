from __future__ import annotations

import contextlib
import math
import os
import posixpath
import re
import zipfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path, PurePosixPath
from typing import Any
from urllib.parse import unquote
from uuid import uuid4
from xml.etree import ElementTree

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

from sts.domain import DomainError, ErrorCode

_GIB = 1024**3
_MIB = 1024**2
_CELL_REFERENCE = re.compile(r"^([A-Za-z]+)([1-9][0-9]*)$")
_RELATIONSHIP_ID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"


@dataclass(frozen=True, slots=True)
class XlsxLimits:
    max_upload_bytes: int = 8 * _GIB
    max_zip_members: int = 4_096
    max_total_uncompressed_bytes: int = 16 * _GIB
    max_member_uncompressed_bytes: int = 8 * _GIB
    max_compression_ratio: float = 100.0
    max_worksheets: int = 64
    max_rows_per_sheet: int = 1_048_576
    max_columns: int = 70
    max_cells_per_sheet: int = 75_000_000
    max_shared_strings_bytes: int = 512 * _MIB

    def __post_init__(self) -> None:
        for field_name in (
            "max_upload_bytes",
            "max_zip_members",
            "max_total_uncompressed_bytes",
            "max_member_uncompressed_bytes",
            "max_worksheets",
            "max_rows_per_sheet",
            "max_columns",
            "max_cells_per_sheet",
            "max_shared_strings_bytes",
        ):
            if getattr(self, field_name) <= 0:
                raise ValueError(f"{field_name} must be positive")
        if not math.isfinite(self.max_compression_ratio) or self.max_compression_ratio <= 0:
            raise ValueError("max_compression_ratio must be finite and positive")


DEFAULT_XLSX_LIMITS = XlsxLimits()


@dataclass(frozen=True, slots=True)
class XlsxSheet:
    name: str
    index: int
    sheet_id: str
    visibility: str
    package_path: str
    row_count: int
    column_count: int
    cell_count: int


@dataclass(frozen=True, slots=True)
class XlsxInspection:
    source_size_bytes: int
    zip_member_count: int
    total_uncompressed_bytes: int
    sheets: tuple[XlsxSheet, ...]

    @property
    def requires_sheet_selection(self) -> bool:
        return len(self.sheets) > 1

    @property
    def automatically_selected_sheet(self) -> str | None:
        return self.sheets[0].name if len(self.sheets) == 1 else None


@dataclass(frozen=True, slots=True)
class XlsxConversionResult:
    path: Path
    selected_sheet: str
    row_count: int
    column_count: int
    record_batch_count: int
    size_bytes: int


@dataclass(frozen=True, slots=True)
class _SheetStats:
    row_count: int
    column_count: int
    cell_count: int


@dataclass(frozen=True, slots=True)
class _WorkbookSheet:
    name: str
    sheet_id: str
    visibility: str
    relationship_id: str


def _unsafe(reason: str, detail: str, **context: Any) -> DomainError:
    return DomainError(
        ErrorCode.XLSX_UNSAFE,
        detail,
        context={"reason": reason, **context},
    )


def _local_name(tag: str) -> str:
    return tag.rpartition("}")[2]


def _validate_member_name(name: str) -> None:
    if not name or "\x00" in name or "\\" in name or name.startswith("/"):
        raise _unsafe("PATH_TRAVERSAL", "XLSX ZIP contains an unsafe member path", member=name)
    trimmed = name[:-1] if name.endswith("/") else name
    parts = trimmed.split("/")
    if (
        not trimmed
        or any(part in {"", ".", ".."} for part in parts)
        or parts[0].endswith(":")
        or str(PurePosixPath(trimmed)) != trimmed
    ):
        raise _unsafe("PATH_TRAVERSAL", "XLSX ZIP contains an unsafe member path", member=name)


def _compression_ratio(info: zipfile.ZipInfo) -> float:
    if info.file_size == 0:
        return 0.0
    if info.compress_size == 0:
        return math.inf
    return info.file_size / info.compress_size


def _column_index(reference: str) -> tuple[int, int]:
    match = _CELL_REFERENCE.fullmatch(reference)
    if match is None:
        raise _unsafe(
            "PACKAGE_STRUCTURE",
            "worksheet contains an invalid cell reference",
            cell_reference=reference,
        )
    column_letters, row_text = match.groups()
    column = 0
    for character in column_letters.upper():
        column = column * 26 + ord(character) - ord("A") + 1
    return column, int(row_text)


def _relationship_target(base_path: str, target: str) -> str:
    decoded = unquote(target)
    if not decoded or "\x00" in decoded or "\\" in decoded:
        raise _unsafe(
            "PATH_TRAVERSAL",
            "XLSX relationship contains an unsafe target path",
            target=target,
        )
    if decoded.startswith("/"):
        candidate = decoded.lstrip("/")
    else:
        candidate = posixpath.join(posixpath.dirname(base_path), decoded)
    normalized = posixpath.normpath(candidate)
    if normalized in {"", ".", ".."} or normalized.startswith("../"):
        raise _unsafe(
            "PATH_TRAVERSAL",
            "XLSX relationship escapes the package root",
            target=target,
        )
    return normalized


def _parse_relationships(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    *,
    source_part: str,
) -> dict[str, tuple[str, str]]:
    relationships: dict[str, tuple[str, str]] = {}
    with archive.open(info) as stream:
        for event, element in ElementTree.iterparse(stream, events=("start", "end")):
            if event == "start" and _local_name(element.tag) == "Relationship":
                relationship_id = element.attrib.get("Id")
                target = element.attrib.get("Target")
                relationship_type = element.attrib.get("Type", "")
                target_mode = element.attrib.get("TargetMode", "")
                if target_mode.casefold() == "external" or relationship_type.casefold().endswith(
                    "/externallink"
                ):
                    raise _unsafe(
                        "EXTERNAL_LINK",
                        "XLSX external relationships are not permitted",
                        relationship_id=relationship_id,
                    )
                if relationship_id and target:
                    if relationship_id in relationships:
                        raise _unsafe(
                            "PACKAGE_STRUCTURE",
                            "XLSX relationship identifiers must be unique",
                            relationship_id=relationship_id,
                        )
                    relationships[relationship_id] = (
                        relationship_type,
                        _relationship_target(source_part, target),
                    )
            if event == "end":
                element.clear()
    return relationships


def _read_workbook_sheets(
    archive: zipfile.ZipFile,
    workbook_info: zipfile.ZipInfo,
) -> tuple[_WorkbookSheet, ...]:
    sheets: list[_WorkbookSheet] = []
    names: set[str] = set()
    with archive.open(workbook_info) as stream:
        for event, element in ElementTree.iterparse(stream, events=("start", "end")):
            if event == "start" and _local_name(element.tag) == "sheet":
                name = element.attrib.get("name")
                relationship_id = element.attrib.get(_RELATIONSHIP_ID)
                sheet_id = element.attrib.get("sheetId")
                if not name or not relationship_id or not sheet_id:
                    raise _unsafe(
                        "PACKAGE_STRUCTURE",
                        "workbook sheet metadata is incomplete",
                    )
                folded_name = name.casefold()
                if folded_name in names:
                    raise _unsafe(
                        "DUPLICATE_SHEET",
                        "workbook sheet names must be unique",
                        sheet=name,
                    )
                names.add(folded_name)
                sheets.append(
                    _WorkbookSheet(
                        name=name,
                        sheet_id=sheet_id,
                        visibility=element.attrib.get("state", "visible"),
                        relationship_id=relationship_id,
                    )
                )
            if event == "end":
                element.clear()
    return tuple(sheets)


def _scan_content_types(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> None:
    with archive.open(info) as stream:
        for _, element in ElementTree.iterparse(stream, events=("end",)):
            content_type = element.attrib.get("ContentType", "").casefold()
            if "macroenabled" in content_type or "vbaproject" in content_type:
                raise _unsafe("MACRO", "macro-enabled XLSX packages are not permitted")
            element.clear()


def _scan_worksheet(archive: zipfile.ZipFile, info: zipfile.ZipInfo) -> _SheetStats:
    row_elements = 0
    max_row = 0
    max_column = 0
    current_row = 0
    next_column = 1

    with archive.open(info) as stream:
        for event, element in ElementTree.iterparse(stream, events=("start", "end")):
            local_name = _local_name(element.tag)
            if event == "start":
                if local_name == "row":
                    row_elements += 1
                    row_reference = element.attrib.get("r")
                    if row_reference is None:
                        current_row += 1
                    else:
                        try:
                            current_row = int(row_reference)
                        except ValueError as error:
                            raise _unsafe(
                                "PACKAGE_STRUCTURE",
                                "worksheet contains an invalid row reference",
                                row_reference=row_reference,
                            ) from error
                        if current_row <= 0:
                            raise _unsafe(
                                "PACKAGE_STRUCTURE",
                                "worksheet contains an invalid row reference",
                                row_reference=row_reference,
                            )
                    next_column = 1
                    max_row = max(max_row, current_row)
                elif local_name == "c":
                    cell_reference = element.attrib.get("r")
                    if cell_reference is None:
                        if current_row <= 0:
                            raise _unsafe(
                                "PACKAGE_STRUCTURE",
                                "worksheet cell has no row context",
                            )
                        column = next_column
                        row = current_row
                    else:
                        column, row = _column_index(cell_reference)
                    next_column = column + 1
                    max_column = max(max_column, column)
                    max_row = max(max_row, row)
                elif local_name == "f":
                    raise _unsafe(
                        "FORMULA",
                        "formula cells are not permitted in XLSX input",
                        worksheet=info.filename,
                    )
            else:
                element.clear()
    row_count = max(max_row, row_elements)
    return _SheetStats(
        row_count=row_count,
        column_count=max_column,
        cell_count=row_count * max_column,
    )


def _enforce_sheet_limits(sheet: XlsxSheet, limits: XlsxLimits) -> None:
    if sheet.row_count > limits.max_rows_per_sheet:
        raise _unsafe(
            "ROW_LIMIT",
            "selected worksheet exceeds the row limit",
            sheet=sheet.name,
            actual=sheet.row_count,
            limit=limits.max_rows_per_sheet,
        )
    if sheet.column_count > limits.max_columns:
        raise _unsafe(
            "COLUMN_LIMIT",
            "selected worksheet exceeds the column limit",
            sheet=sheet.name,
            actual=sheet.column_count,
            limit=limits.max_columns,
        )
    if sheet.cell_count > limits.max_cells_per_sheet:
        raise _unsafe(
            "CELL_LIMIT",
            "selected worksheet exceeds the populated-cell limit",
            sheet=sheet.name,
            actual=sheet.cell_count,
            limit=limits.max_cells_per_sheet,
        )


def preflight_xlsx(
    source_path: str | os.PathLike[str],
    *,
    limits: XlsxLimits = DEFAULT_XLSX_LIMITS,
) -> XlsxInspection:
    """Validate an XLSX package without invoking OpenPyXL or materializing XML trees."""

    source = Path(source_path)
    try:
        source_size = source.stat().st_size
    except OSError as error:
        raise _unsafe("PACKAGE_STRUCTURE", "XLSX input is not a readable regular file") from error
    if not source.is_file():
        raise _unsafe("PACKAGE_STRUCTURE", "XLSX input is not a regular file")
    if source_size > limits.max_upload_bytes:
        raise DomainError(
            ErrorCode.UPLOAD_TOO_LARGE,
            "XLSX upload exceeds the configured upload limit",
            context={"actual": source_size, "limit": limits.max_upload_bytes},
        )

    try:
        with zipfile.ZipFile(source, mode="r") as archive:
            members = archive.infolist()
            if len(members) > limits.max_zip_members:
                raise _unsafe(
                    "MEMBER_COUNT_LIMIT",
                    "XLSX ZIP contains too many members",
                    actual=len(members),
                    limit=limits.max_zip_members,
                )

            by_name: dict[str, zipfile.ZipInfo] = {}
            total_uncompressed = 0
            for info in members:
                _validate_member_name(info.filename)
                if info.filename in by_name:
                    raise _unsafe(
                        "DUPLICATE_MEMBER",
                        "XLSX ZIP contains duplicate member names",
                        member=info.filename,
                    )
                by_name[info.filename] = info
                if info.flag_bits & 0x1:
                    raise _unsafe(
                        "ENCRYPTED_MEMBER",
                        "encrypted XLSX ZIP members are not permitted",
                        member=info.filename,
                    )
                if info.file_size > limits.max_member_uncompressed_bytes:
                    raise _unsafe(
                        "MEMBER_UNCOMPRESSED_LIMIT",
                        "XLSX ZIP member exceeds the uncompressed-size limit",
                        member=info.filename,
                        actual=info.file_size,
                        limit=limits.max_member_uncompressed_bytes,
                    )
                ratio = _compression_ratio(info)
                if ratio > limits.max_compression_ratio:
                    raise _unsafe(
                        "COMPRESSION_RATIO_LIMIT",
                        "XLSX ZIP member exceeds the compression-ratio limit",
                        member=info.filename,
                        actual=ratio,
                        limit=limits.max_compression_ratio,
                    )
                total_uncompressed += info.file_size
                if total_uncompressed > limits.max_total_uncompressed_bytes:
                    raise _unsafe(
                        "TOTAL_UNCOMPRESSED_LIMIT",
                        "XLSX ZIP exceeds the total uncompressed-size limit",
                        actual=total_uncompressed,
                        limit=limits.max_total_uncompressed_bytes,
                    )

                member_name = info.filename.casefold()
                if member_name == "xl/sharedstrings.xml" and (
                    info.file_size > limits.max_shared_strings_bytes
                ):
                    raise _unsafe(
                        "SHARED_STRINGS_LIMIT",
                        "XLSX shared strings XML exceeds the size limit",
                        actual=info.file_size,
                        limit=limits.max_shared_strings_bytes,
                    )
                if (
                    member_name == "xl/vbaproject.bin"
                    or member_name.startswith("xl/macrosheets/")
                    or member_name.startswith("xl/dialogsheets/")
                ):
                    raise _unsafe("MACRO", "macro-enabled XLSX packages are not permitted")
                if member_name.startswith("xl/externallinks/"):
                    raise _unsafe(
                        "EXTERNAL_LINK",
                        "XLSX external links are not permitted",
                        member=info.filename,
                    )

            content_types_info = by_name.get("[Content_Types].xml")
            workbook_info = by_name.get("xl/workbook.xml")
            workbook_relationships_info = by_name.get("xl/_rels/workbook.xml.rels")
            if (
                content_types_info is None
                or workbook_info is None
                or workbook_relationships_info is None
            ):
                raise _unsafe(
                    "PACKAGE_STRUCTURE",
                    "XLSX package is missing required workbook metadata",
                )

            _scan_content_types(archive, content_types_info)
            for info in members:
                if info.filename.endswith(".rels"):
                    _parse_relationships(archive, info, source_part=info.filename)

            relationships = _parse_relationships(
                archive,
                workbook_relationships_info,
                source_part="xl/workbook.xml",
            )
            workbook_sheets = _read_workbook_sheets(archive, workbook_info)
            worksheet_bindings: list[tuple[_WorkbookSheet, str]] = []
            for workbook_sheet in workbook_sheets:
                relationship = relationships.get(workbook_sheet.relationship_id)
                if relationship is None:
                    raise _unsafe(
                        "PACKAGE_STRUCTURE",
                        "worksheet relationship is missing",
                        sheet=workbook_sheet.name,
                    )
                relationship_type, package_path = relationship
                if not relationship_type.casefold().endswith("/worksheet"):
                    continue
                if package_path not in by_name:
                    raise _unsafe(
                        "PACKAGE_STRUCTURE",
                        "worksheet package member is missing",
                        sheet=workbook_sheet.name,
                        member=package_path,
                    )
                worksheet_bindings.append((workbook_sheet, package_path))

            if not worksheet_bindings:
                raise _unsafe("PACKAGE_STRUCTURE", "XLSX workbook contains no worksheets")
            if len(worksheet_bindings) > limits.max_worksheets:
                raise _unsafe(
                    "WORKSHEET_LIMIT",
                    "XLSX workbook contains too many worksheets",
                    actual=len(worksheet_bindings),
                    limit=limits.max_worksheets,
                )

            stats_by_path: dict[str, _SheetStats] = {}
            for info in members:
                member_name = info.filename.casefold()
                if member_name.startswith("xl/worksheets/") and member_name.endswith(".xml"):
                    stats_by_path[info.filename] = _scan_worksheet(archive, info)

            sheets: list[XlsxSheet] = []
            for index, (workbook_sheet, package_path) in enumerate(worksheet_bindings):
                stats = stats_by_path.get(package_path)
                if stats is None:
                    raise _unsafe(
                        "PACKAGE_STRUCTURE",
                        "worksheet metadata could not be inspected",
                        sheet=workbook_sheet.name,
                    )
                sheets.append(
                    XlsxSheet(
                        name=workbook_sheet.name,
                        index=index,
                        sheet_id=workbook_sheet.sheet_id,
                        visibility=workbook_sheet.visibility,
                        package_path=package_path,
                        row_count=stats.row_count,
                        column_count=stats.column_count,
                        cell_count=stats.cell_count,
                    )
                )

            inspection = XlsxInspection(
                source_size_bytes=source_size,
                zip_member_count=len(members),
                total_uncompressed_bytes=total_uncompressed,
                sheets=tuple(sheets),
            )
            if len(inspection.sheets) == 1:
                _enforce_sheet_limits(inspection.sheets[0], limits)
            return inspection
    except DomainError:
        raise
    except (ElementTree.ParseError, KeyError, OSError, ValueError, zipfile.BadZipFile) as error:
        raise _unsafe("PACKAGE_STRUCTURE", "XLSX package is invalid or corrupt") from error


def select_xlsx_sheet(
    inspection: XlsxInspection,
    selected_sheet: str | Sequence[str] | None,
    *,
    limits: XlsxLimits = DEFAULT_XLSX_LIMITS,
    concatenate_sheets: bool = False,
) -> XlsxSheet:
    if concatenate_sheets or (
        isinstance(selected_sheet, Sequence) and not isinstance(selected_sheet, str)
    ):
        raise DomainError(
            ErrorCode.INPUT_FORMAT_UNSUPPORTED,
            "multi-sheet XLSX concatenation is not supported",
        )
    if selected_sheet is None:
        if inspection.requires_sheet_selection:
            raise DomainError(
                ErrorCode.INVALID_STATE,
                "a single worksheet must be selected before XLSX conversion",
                context={
                    "required_state": "sheet_required",
                    "available_sheets": [sheet.name for sheet in inspection.sheets],
                },
            )
        sheet = inspection.sheets[0]
    else:
        matches = [sheet for sheet in inspection.sheets if sheet.name == selected_sheet]
        if not matches:
            raise DomainError(
                ErrorCode.INVALID_STATE,
                "selected worksheet does not exist in the inspected workbook",
                context={
                    "selected_sheet": selected_sheet,
                    "available_sheets": [sheet.name for sheet in inspection.sheets],
                },
            )
        sheet = matches[0]
    _enforce_sheet_limits(sheet, limits)
    return sheet


def _raw_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, float):
        if not math.isfinite(value):
            raise DomainError(
                ErrorCode.SCHEMA_INVALID,
                "non-finite XLSX numeric values cannot be represented in raw input",
            )
        return str(value)
    return str(value)


def _headers(values: tuple[Any, ...]) -> tuple[str, ...]:
    headers: list[str] = []
    seen: set[str] = set()
    for column_index, value in enumerate(values, start=1):
        header = _raw_text(value)
        if header is None or not header.strip():
            raise DomainError(
                ErrorCode.SCHEMA_INVALID,
                "XLSX header cells must be non-empty",
                context={"column": column_index},
            )
        if header == "__sts_row_id":
            raise DomainError(
                ErrorCode.SCHEMA_INVALID,
                "XLSX header uses the reserved __sts_row_id column name",
            )
        if header in seen:
            raise DomainError(
                ErrorCode.SCHEMA_INVALID,
                "XLSX header names must be unique",
                context={"column": column_index, "header": header},
            )
        seen.add(header)
        headers.append(header)
    return tuple(headers)


def _record_batch(
    row_ids: list[int],
    columns: list[list[str | None]],
    schema: pa.Schema,
) -> pa.RecordBatch:
    arrays: list[pa.Array] = [pa.array(row_ids, type=pa.int64())]
    arrays.extend(pa.array(values, type=pa.string()) for values in columns)
    return pa.RecordBatch.from_arrays(arrays, schema=schema)


def _fsync_file(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY)
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def _fsync_directory(path: Path) -> None:
    descriptor = os.open(path, os.O_RDONLY | getattr(os, "O_DIRECTORY", 0))
    try:
        os.fsync(descriptor)
    finally:
        os.close(descriptor)


def convert_xlsx_to_raw_parquet(
    source_path: str | os.PathLike[str],
    output_path: str | os.PathLike[str],
    *,
    selected_sheet: str | Sequence[str] | None = None,
    concatenate_sheets: bool = False,
    output_format: str = "parquet",
    limits: XlsxLimits = DEFAULT_XLSX_LIMITS,
    batch_rows: int = 65_536,
) -> XlsxConversionResult:
    """Stream one preflighted worksheet into an atomically published raw Parquet file."""

    if output_format.casefold() != "parquet" or Path(output_path).suffix.casefold() == ".xlsx":
        raise DomainError(
            ErrorCode.INPUT_FORMAT_UNSUPPORTED,
            "XLSX output is not supported; raw ingest output must be Parquet",
        )
    if batch_rows <= 0:
        raise ValueError("batch_rows must be positive")

    inspection = preflight_xlsx(source_path, limits=limits)
    sheet = select_xlsx_sheet(
        inspection,
        selected_sheet,
        limits=limits,
        concatenate_sheets=concatenate_sheets,
    )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists() or output.is_symlink():
        raise DomainError(
            ErrorCode.IMMUTABLE_PATH_EXISTS,
            f"immutable raw Parquet path already exists: {output}",
        )
    part = output.with_name(f".{output.name}.{uuid4().hex}.part")
    workbook = None
    writer: pq.ParquetWriter | None = None
    published = False
    try:
        workbook = load_workbook(
            filename=Path(source_path),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        worksheet = workbook[sheet.name]
        rows = worksheet.iter_rows(
            min_row=1,
            max_row=sheet.row_count,
            min_col=1,
            max_col=sheet.column_count,
            values_only=False,
        )
        try:
            header_cells = next(rows)
        except StopIteration as error:
            raise DomainError(
                ErrorCode.SCHEMA_INVALID, "selected XLSX worksheet is empty"
            ) from error
        for cell in header_cells:
            if cell.data_type == "f":
                raise _unsafe("FORMULA", "formula cells are not permitted in XLSX input")
        headers = _headers(tuple(cell.value for cell in header_cells))
        schema = pa.schema(
            [pa.field("__sts_row_id", pa.int64(), nullable=False)]
            + [pa.field(header, pa.string()) for header in headers]
        )
        writer = pq.ParquetWriter(
            part,
            schema,
            compression="zstd",
            use_dictionary=False,
            write_statistics=True,
        )
        os.chmod(part, 0o600)

        row_ids: list[int] = []
        columns: list[list[str | None]] = [[] for _ in headers]
        row_count = 0
        record_batch_count = 0
        for cells in rows:
            if len(cells) != len(headers):
                raise _unsafe(
                    "PACKAGE_STRUCTURE",
                    "worksheet row width changed during streaming conversion",
                )
            values: list[str | None] = []
            for cell in cells:
                if cell.data_type == "f":
                    raise _unsafe("FORMULA", "formula cells are not permitted in XLSX input")
                values.append(_raw_text(cell.value))
            row_ids.append(row_count)
            for column, value in zip(columns, values, strict=True):
                column.append(value)
            row_count += 1
            if len(row_ids) >= batch_rows:
                writer.write_batch(_record_batch(row_ids, columns, schema))
                record_batch_count += 1
                row_ids = []
                columns = [[] for _ in headers]
        if row_ids:
            writer.write_batch(_record_batch(row_ids, columns, schema))
            record_batch_count += 1
        writer.close()
        writer = None
        _fsync_file(part)
        try:
            os.link(part, output)
        except FileExistsError as error:
            raise DomainError(
                ErrorCode.IMMUTABLE_PATH_EXISTS,
                f"immutable raw Parquet path already exists: {output}",
            ) from error
        published = True
        part.unlink()
        _fsync_directory(output.parent)
        return XlsxConversionResult(
            path=output,
            selected_sheet=sheet.name,
            row_count=row_count,
            column_count=len(headers),
            record_batch_count=record_batch_count,
            size_bytes=output.stat().st_size,
        )
    finally:
        if writer is not None:
            writer.close()
        if workbook is not None:
            workbook.close()
        if not published:
            with contextlib.suppress(FileNotFoundError):
                part.unlink()
