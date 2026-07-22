from __future__ import annotations

import os
import struct
import zipfile
from dataclasses import replace
from pathlib import Path
from typing import Any

import openpyxl
import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from openpyxl import Workbook

from sts.domain import DomainError, ErrorCode
from sts.ingest import xlsx
from sts.ingest.xlsx import (
    DEFAULT_XLSX_LIMITS,
    XlsxLimits,
    convert_xlsx_to_raw_parquet,
    preflight_xlsx,
    select_xlsx_sheet,
)


def _workbook(path: Path, sheets: dict[str, list[list[Any]]]) -> Path:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    for name, rows in sheets.items():
        worksheet = workbook.create_sheet(name)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    workbook.close()
    return path


def _append_member(
    path: Path,
    name: str,
    content: bytes,
    *,
    compression: int = zipfile.ZIP_DEFLATED,
) -> None:
    with zipfile.ZipFile(path, mode="a", compression=compression) as archive:
        archive.writestr(name, content, compress_type=compression)


def _mark_first_member_encrypted(path: Path) -> None:
    payload = bytearray(path.read_bytes())
    central_header = payload.index(b"PK\x01\x02")
    flag_offset = central_header + 8
    flags = struct.unpack_from("<H", payload, flag_offset)[0]
    struct.pack_into("<H", payload, flag_offset, flags | 0x1)
    path.write_bytes(payload)


def _assert_unsafe(error: pytest.ExceptionInfo[DomainError], reason: str) -> None:
    assert error.value.code == ErrorCode.XLSX_UNSAFE
    assert error.value.status == 422
    assert error.value.problem.context["reason"] == reason


def test_safe_single_sheet_conversion_streams_zstd_arrow_batches(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = _workbook(
        tmp_path / "safe.xlsx",
        {
            "Data": [
                ["name", "amount", "active", "missing"],
                ["Alice", 10, True, None],
                ["Bob", 12.5, False, ""],
                ["Carol", None, None, "value"],
            ]
        },
    )
    output = tmp_path / "raw.parquet"
    real_loader = openpyxl.load_workbook
    loader_calls: list[dict[str, Any]] = []
    selected_worksheets: list[str] = []

    class WorkbookProxy:
        def __init__(self, workbook: Any) -> None:
            self._workbook = workbook

        def __getitem__(self, name: str) -> Any:
            selected_worksheets.append(name)
            return self._workbook[name]

        def close(self) -> None:
            self._workbook.close()

    def tracking_loader(*args: Any, **kwargs: Any) -> WorkbookProxy:
        loader_calls.append(kwargs)
        return WorkbookProxy(real_loader(*args, **kwargs))

    monkeypatch.setattr(xlsx, "load_workbook", tracking_loader)
    result = convert_xlsx_to_raw_parquet(source, output, batch_rows=1)

    assert loader_calls == [
        {
            "filename": source,
            "read_only": True,
            "data_only": False,
            "keep_links": False,
        }
    ]
    assert selected_worksheets == ["Data"]
    assert result.selected_sheet == "Data"
    assert result.row_count == 3
    assert result.column_count == 4
    assert result.record_batch_count == 3
    assert result.size_bytes == output.stat().st_size
    assert not list(tmp_path.glob("*.part"))
    assert not list(tmp_path.glob(".*.part"))

    parquet = pq.ParquetFile(output)
    assert parquet.metadata.num_row_groups == 3
    for row_group in range(parquet.metadata.num_row_groups):
        for column in range(parquet.metadata.num_columns):
            assert (
                parquet.metadata.row_group(row_group).column(column).compression
                == "ZSTD"
            )

    table = parquet.read()
    assert table.schema.names == ["__sts_row_id", "name", "amount", "active", "missing"]
    assert table.schema.field("__sts_row_id").type == pa.int64()
    assert all(
        pa.types.is_string(table.schema.field(name).type)
        for name in table.schema.names[1:]
    )
    assert table.column("__sts_row_id").to_pylist() == [0, 1, 2]
    assert table.column("name").to_pylist() == ["Alice", "Bob", "Carol"]
    assert table.column("amount").to_pylist() == ["10", "12.5", None]
    assert table.column("active").to_pylist() == ["true", "false", None]
    assert table.column("missing").to_pylist() == [None, None, "value"]


def test_preflight_lists_sheets_and_requires_one_explicit_selection(
    tmp_path: Path,
) -> None:
    source = _workbook(
        tmp_path / "multiple.xlsx",
        {
            "First": [["value"], [1]],
            "Second": [["value"], [2]],
        },
    )
    workbook = openpyxl.load_workbook(source)
    workbook["Second"].sheet_state = "hidden"
    workbook.save(source)
    workbook.close()

    inspection = preflight_xlsx(source)

    assert inspection.requires_sheet_selection is True
    assert inspection.automatically_selected_sheet is None
    assert [
        (sheet.name, sheet.index, sheet.visibility) for sheet in inspection.sheets
    ] == [
        ("First", 0, "visible"),
        ("Second", 1, "hidden"),
    ]
    with pytest.raises(DomainError) as required:
        select_xlsx_sheet(inspection, None)
    assert required.value.code == ErrorCode.INVALID_STATE
    assert required.value.problem.context["required_state"] == "sheet_required"
    assert required.value.problem.context["available_sheets"] == ["First", "Second"]

    selected = select_xlsx_sheet(inspection, "Second")
    assert selected.name == "Second"
    output = tmp_path / "selected.parquet"
    result = convert_xlsx_to_raw_parquet(source, output, selected_sheet="Second")
    assert result.row_count == 1
    assert pq.read_table(output).column("value").to_pylist() == ["2"]


def test_selected_sheet_limits_do_not_reject_an_unselected_sheet(
    tmp_path: Path,
) -> None:
    source = _workbook(
        tmp_path / "selected-limit.xlsx",
        {
            "Small": [["value"], [1]],
            "Large": [["value"], [1], [2], [3]],
        },
    )
    limits = replace(DEFAULT_XLSX_LIMITS, max_rows_per_sheet=2)
    inspection = preflight_xlsx(source, limits=limits)

    assert select_xlsx_sheet(inspection, "Small", limits=limits).name == "Small"
    with pytest.raises(DomainError) as error:
        select_xlsx_sheet(inspection, "Large", limits=limits)
    _assert_unsafe(error, "ROW_LIMIT")


def test_multi_sheet_concatenation_and_xlsx_output_are_rejected(tmp_path: Path) -> None:
    source = _workbook(
        tmp_path / "unsupported.xlsx",
        {"First": [["value"], [1]], "Second": [["value"], [2]]},
    )
    inspection = preflight_xlsx(source)

    with pytest.raises(DomainError) as sequence_error:
        select_xlsx_sheet(inspection, ["First", "Second"])
    assert sequence_error.value.code == ErrorCode.INPUT_FORMAT_UNSUPPORTED
    with pytest.raises(DomainError) as concatenate_error:
        convert_xlsx_to_raw_parquet(
            source,
            tmp_path / "concatenated.parquet",
            selected_sheet="First",
            concatenate_sheets=True,
        )
    assert concatenate_error.value.code == ErrorCode.INPUT_FORMAT_UNSUPPORTED
    with pytest.raises(DomainError) as output_error:
        convert_xlsx_to_raw_parquet(
            source,
            tmp_path / "output.xlsx",
            selected_sheet="First",
            output_format="xlsx",
        )
    assert output_error.value.code == ErrorCode.INPUT_FORMAT_UNSUPPORTED


def test_zip_member_count_limit_is_literal_and_injectable(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "members.xlsx", {"Data": [["value"], [1]]})
    with zipfile.ZipFile(source) as archive:
        actual_members = len(archive.infolist())
    limits = replace(DEFAULT_XLSX_LIMITS, max_zip_members=actual_members - 1)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    _assert_unsafe(error, "MEMBER_COUNT_LIMIT")


def test_total_uncompressed_size_limit_is_literal_and_injectable(
    tmp_path: Path,
) -> None:
    source = _workbook(tmp_path / "total-size.xlsx", {"Data": [["value"], [1]]})
    limits = replace(DEFAULT_XLSX_LIMITS, max_total_uncompressed_bytes=1)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    _assert_unsafe(error, "TOTAL_UNCOMPRESSED_LIMIT")


def test_single_member_size_limit_is_literal_and_injectable(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "member-size.xlsx", {"Data": [["value"], [1]]})
    limits = replace(DEFAULT_XLSX_LIMITS, max_member_uncompressed_bytes=1)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    _assert_unsafe(error, "MEMBER_UNCOMPRESSED_LIMIT")


def test_compression_ratio_limit_is_literal_and_injectable(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "ratio.xlsx", {"Data": [["value"], [1]]})
    _append_member(source, "padding.bin", b"0" * 100_000)
    limits = replace(DEFAULT_XLSX_LIMITS, max_compression_ratio=10.0)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    _assert_unsafe(error, "COMPRESSION_RATIO_LIMIT")


def test_worksheet_count_limit_is_literal_and_injectable(tmp_path: Path) -> None:
    source = _workbook(
        tmp_path / "sheet-count.xlsx",
        {"First": [["value"]], "Second": [["value"]]},
    )
    limits = replace(DEFAULT_XLSX_LIMITS, max_worksheets=1)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    _assert_unsafe(error, "WORKSHEET_LIMIT")


@pytest.mark.parametrize(
    ("limit_name", "limit_value", "reason"),
    [
        ("max_rows_per_sheet", 2, "ROW_LIMIT"),
        ("max_columns", 2, "COLUMN_LIMIT"),
        ("max_cells_per_sheet", 5, "CELL_LIMIT"),
    ],
)
def test_selected_sheet_shape_limits_are_literal_and_injectable(
    tmp_path: Path,
    limit_name: str,
    limit_value: int,
    reason: str,
) -> None:
    source = _workbook(
        tmp_path / f"{limit_name}.xlsx",
        {"Data": [["a", "b", "c"], [1, 2, 3], [4, 5, 6]]},
    )
    limits = replace(DEFAULT_XLSX_LIMITS, **{limit_name: limit_value})

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    _assert_unsafe(error, reason)


def test_shared_strings_limit_is_literal_and_injectable(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "shared-strings.xlsx", {"Data": [["value"], [1]]})
    _append_member(source, "xl/sharedStrings.xml", b"x" * 64)
    limits = replace(DEFAULT_XLSX_LIMITS, max_shared_strings_bytes=32)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    _assert_unsafe(error, "SHARED_STRINGS_LIMIT")


def test_upload_limit_is_enforced_before_zip_open(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "upload-size.xlsx", {"Data": [["value"], [1]]})
    limits = replace(DEFAULT_XLSX_LIMITS, max_upload_bytes=source.stat().st_size - 1)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source, limits=limits)
    assert error.value.code == ErrorCode.UPLOAD_TOO_LARGE
    assert error.value.status == 413


def test_zip_path_traversal_is_rejected(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "traversal.xlsx", {"Data": [["value"], [1]]})
    _append_member(source, "../outside", b"private")

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source)
    _assert_unsafe(error, "PATH_TRAVERSAL")


def test_duplicate_zip_member_names_are_rejected(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "duplicate.xlsx", {"Data": [["value"], [1]]})
    with zipfile.ZipFile(source) as archive:
        content_types = archive.read("[Content_Types].xml")
    with pytest.warns(UserWarning, match="Duplicate name"):
        _append_member(source, "[Content_Types].xml", content_types)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source)
    _assert_unsafe(error, "DUPLICATE_MEMBER")


def test_encrypted_zip_member_is_rejected_from_central_directory(
    tmp_path: Path,
) -> None:
    source = _workbook(tmp_path / "encrypted.xlsx", {"Data": [["value"], [1]]})
    _mark_first_member_encrypted(source)

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source)
    _assert_unsafe(error, "ENCRYPTED_MEMBER")


def test_macro_member_is_rejected(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "macro.xlsx", {"Data": [["value"], [1]]})
    _append_member(source, "xl/vbaProject.bin", b"macro")

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source)
    _assert_unsafe(error, "MACRO")


def test_external_link_member_is_rejected(tmp_path: Path) -> None:
    source = _workbook(tmp_path / "external.xlsx", {"Data": [["value"], [1]]})
    _append_member(source, "xl/externalLinks/externalLink1.xml", b"<externalLink/>")

    with pytest.raises(DomainError) as error:
        preflight_xlsx(source)
    _assert_unsafe(error, "EXTERNAL_LINK")


def test_formula_cell_is_rejected_before_openpyxl_is_invoked(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    source = _workbook(
        tmp_path / "formula.xlsx",
        {"Data": [["left", "right", "sum"], [1, 2, "=A2+B2"]]},
    )
    loader_called = False

    def forbidden_loader(*args: Any, **kwargs: Any) -> None:
        nonlocal loader_called
        loader_called = True
        raise AssertionError(
            "OpenPyXL must not run before ZIP preflight accepts the workbook"
        )

    monkeypatch.setattr(xlsx, "load_workbook", forbidden_loader)
    with pytest.raises(DomainError) as error:
        convert_xlsx_to_raw_parquet(source, tmp_path / "formula.parquet")
    _assert_unsafe(error, "FORMULA")
    assert loader_called is False
    assert not (tmp_path / "formula.parquet").exists()


def test_default_limits_match_the_ingest_contract() -> None:
    assert XlsxLimits() == DEFAULT_XLSX_LIMITS
    assert DEFAULT_XLSX_LIMITS.max_upload_bytes == 8 * 1024**3
    assert DEFAULT_XLSX_LIMITS.max_zip_members == 4_096
    assert DEFAULT_XLSX_LIMITS.max_total_uncompressed_bytes == 16 * 1024**3
    assert DEFAULT_XLSX_LIMITS.max_member_uncompressed_bytes == 8 * 1024**3
    assert DEFAULT_XLSX_LIMITS.max_compression_ratio == 100.0
    assert DEFAULT_XLSX_LIMITS.max_worksheets == 64
    assert DEFAULT_XLSX_LIMITS.max_rows_per_sheet == 1_048_576
    assert DEFAULT_XLSX_LIMITS.max_columns == 70
    assert DEFAULT_XLSX_LIMITS.max_cells_per_sheet == 75_000_000
    assert DEFAULT_XLSX_LIMITS.max_shared_strings_bytes == 512 * 1024**2


def test_failed_conversion_removes_partial_parquet(tmp_path: Path) -> None:
    source = _workbook(
        tmp_path / "bad-header.xlsx",
        {"Data": [["duplicate", "duplicate"], [1, 2]]},
    )
    output = tmp_path / "bad-header.parquet"

    with pytest.raises(DomainError) as error:
        convert_xlsx_to_raw_parquet(source, output)
    assert error.value.code == ErrorCode.SCHEMA_INVALID
    assert not output.exists()
    assert not [path for path in os.scandir(tmp_path) if path.name.endswith(".part")]
